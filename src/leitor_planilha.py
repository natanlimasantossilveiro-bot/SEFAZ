from openpyxl import load_workbook

from src.utils import limpar_documento, identificar_tipo_documento, validar_documento

COLUNA_DOCUMENTO = "documento"


def obter_cabecalhos(aba):
    cabecalhos = {}

    for indice, celula in enumerate(aba[1], start=1):
        if celula.value:
            nome_coluna = str(celula.value).strip().lower()
            cabecalhos[nome_coluna] = indice

    return cabecalhos


def validar_cabecalhos(cabecalhos):
    if COLUNA_DOCUMENTO not in cabecalhos:
        raise ValueError("A planilha deve conter a coluna 'documento'.")


def ler_documentos_planilha(caminho_planilha):

    workbook = load_workbook(caminho_planilha)

    aba = workbook.active

    cabecalhos = obter_cabecalhos(aba)

    validar_cabecalhos(cabecalhos)

    documentos = []

    coluna_documento = cabecalhos[COLUNA_DOCUMENTO]

    for numero_linha in range(2, aba.max_row + 1):
        valor_documento = aba.cell(row=numero_linha, column=coluna_documento).value

        documento_limpo = limpar_documento(valor_documento)

        if not documento_limpo:
            continue

        tipo_documento = identificar_tipo_documento(documento_limpo)

        documento_valido = validar_documento(documento_limpo)

        documentos.append(
            {
                "linha": numero_linha,
                "documento": documento_limpo,
                "tipo": tipo_documento,
                "valido": documento_valido,
            }
        )

    return documentos
